import io
import openpyxl
from src.services.structural_extractor.parsing.xlsx_parser import parse_xlsx
from src.services.structural_extractor.parsing.model import CellRef


def _build_sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws["A1"] = "Invoice No"
    ws["B1"] = "INV-001"
    ws["A2"] = "Date"
    ws["B2"] = "01/07/2022"
    ws["A4"] = "Description"
    ws["B4"] = "Qty"
    ws["C4"] = "Price"
    ws["A5"] = "Widget"
    ws["B5"] = 10
    ws["C5"] = 99.99
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_cellrefs():
    doc = parse_xlsx(_build_sample_xlsx(), "inv.xlsx")
    assert doc.source_format == "xlsx"
    assert doc.pages_or_sheets == 1
    t = next(t for t in doc.tokens if t.text == "INV-001")
    assert isinstance(t.anchor, CellRef)
    assert t.anchor.sheet == "Invoice"
    assert t.anchor.row == 1 and t.anchor.col == 2


def test_parse_xlsx_table_detected():
    doc = parse_xlsx(_build_sample_xlsx(), "inv.xlsx")
    assert len(doc.tables) >= 1
