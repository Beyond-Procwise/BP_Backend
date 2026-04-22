from io import BytesIO
import openpyxl
from src.services.structural_extractor.parsing.model import (
    CellRef, Token, Region, ParsedDocument, Table
)
from src.services.structural_extractor.exceptions import XlsxParseError


def parse_xlsx(file_bytes: bytes, filename: str) -> ParsedDocument:
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise XlsxParseError(f"openpyxl failed for {filename}: {exc}") from exc

    tokens: list[Token] = []
    regions: list[Region] = []
    tables: list[Table] = []
    full_text_parts: list[str] = []
    order = 0

    for ws in wb.worksheets:
        merged_lookup = {}
        for merged in ws.merged_cells.ranges:
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_lookup[(r, c)] = str(merged)

        sheet_regions: list[list[Region]] = []
        non_empty_count = 0
        for row_cells in ws.iter_rows():
            row_regions: list[Region] = []
            for cell in row_cells:
                if cell.value is None:
                    row_regions.append(Region(tokens=[], kind="cell"))
                    continue
                text = str(cell.value)
                non_empty_count += 1
                merged = merged_lookup.get((cell.row, cell.column))
                tok = Token(
                    text=text,
                    anchor=CellRef(
                        sheet=ws.title,
                        row=cell.row,
                        col=cell.column,
                        merged_range=merged,
                    ),
                    order=order,
                )
                order += 1
                tokens.append(tok)
                row_regions.append(Region(tokens=[tok], kind="cell"))
                full_text_parts.append(text)
            sheet_regions.append(row_regions)

        if non_empty_count >= 4:
            header_idx = None
            for r_idx, row in enumerate(sheet_regions):
                non_empty = [r for r in row if r.tokens]
                if non_empty and all(
                    not r.tokens[0].text.replace(".", "").replace("-", "").isdigit()
                    for r in non_empty
                ):
                    header_idx = r_idx
                    break
            tables.append(Table(rows=sheet_regions, header_row_index=header_idx))

        regions.extend([r for row in sheet_regions for r in row])

    return ParsedDocument(
        source_format="xlsx",
        filename=filename,
        tokens=tokens,
        regions=regions,
        tables=tables,
        pages_or_sheets=len(wb.worksheets),
        full_text="\n".join(full_text_parts),
        raw_bytes=file_bytes,
    )
